# -*- coding: utf-8 -*-
__author__ = 'SanQuan'
__date__ = '2020/6/29'
import openpyxl
import os
import logging
import re
from config.config import COMMON_CFG

logging.basicConfig(format='%(asctime)s - %(pathname)s[line:%(lineno)d] ''- %(levelname)s: %(message)s',
                    level=logging.INFO)


# 遍历project
def traverse_project():
    suite_dir = COMMON_CFG.get("PROJECT_PATH")
    projects = list(set(os.listdir(suite_dir)) - set([name for name in os.listdir(suite_dir) if name.startswith("_")]))
    # 执行匹配条件的测试用例
    if COMMON_CFG["CURRENT_PROJECT_NAME"]:
        project_case_dir = os.path.join(os.path.join(suite_dir, COMMON_CFG["CURRENT_PROJECT_NAME"]), "testcase")
        logging.debug("project_case_dir:%s" % project_case_dir)
        traverse_case_dir(project_case_dir)
    else:
        for project in projects:
            project_case_dir = os.path.join(os.path.join(suite_dir, project), "testcase")
            logging.debug("project_case_dir:%s" % project_case_dir)
            traverse_case_dir(project_case_dir)


# 遍历case目录
def traverse_case_dir(project_case_dir):
    for root, dirs, files in os.walk(project_case_dir):
        dirs = [d for d in dirs if not d.startswith("_")]
        files = [f for f in files if not f.startswith("_") and not f.startswith("~$") and f.endswith("xlsx")]
        logging.debug("Root:%s Dirs:%s Files:%s" % (root, dirs, files))
        for file in files:
            case_file_name = os.path.join(root, file)
            logging.debug("case_file_name:%s" % case_file_name)
            read_excel_file(case_file_name)


# 读取Excel
def read_excel_file(case_file_name):
    wb = openpyxl.load_workbook(case_file_name)
    sheets_name = wb.sheetnames
    logging.debug("sheets_name:%s" % sheets_name)
    template_sheet = wb[sheets_name[0]]
    data_sheet = wb[sheets_name[1]]
    case_template = get_case_template(template_sheet)
    cases_value = get_case_value(data_sheet)
    create_py_test(case_file_name, case_template, cases_value)


# 获取template
def get_case_template(template_sheet):
    max_row = template_sheet.max_row
    max_column = template_sheet.max_column
    logging.debug("max_column:%s" % max_row)
    logging.debug("max_column:%s" % max_column)
    for row in range(max_row):
        if template_sheet.cell(row=row + 1, column=1).value == "参数":
            pass
        elif template_sheet.cell(row=row + 1, column=1).value == "开始用例":
            case_template = {}
            parameters = []
            parameter = []
        elif template_sheet.cell(row=row + 1, column=1).value == "结束用例":
            case_template[step] = parameters
            logging.debug("case_template:%s" % case_template)
            return case_template
        elif template_sheet.cell(row=row + 1, column=1).value.lower() == "step":
            if parameters:
                case_template[step] = parameters
            parameters = []
            step = template_sheet.cell(row=row + 1, column=2).value + "|" \
                   + template_sheet.cell(row=row + 1, column=3).value + "|" \
                   + template_sheet.cell(row=row + 1, column=4).value
        else:
            for column in range(max_column):
                parameter.append(template_sheet.cell(row=row + 1, column=column + 1).value)
            if parameter:
                parameters.append(parameter)
                parameter = []


# 获取template下的用例
def get_case_value(data_sheet):
    cases_value = {}
    case_value = {}
    max_row = data_sheet.max_row
    max_column = data_sheet.max_column
    logging.debug("max_column:%s" % max_row)
    logging.debug("max_column:%s" % max_column)
    for column in range(1, max_column):
        if data_sheet.cell(row=1, column=column + 1).value:
            for row in range(max_row):
                if data_sheet.cell(row=row + 1, column=1).value == "值":
                    case_name = data_sheet.cell(row=row + 1, column=column + 1).value
                elif data_sheet.cell(row=row + 1, column=1).value:
                    key = data_sheet.cell(row=row + 1, column=1).value
                    value = data_sheet.cell(row=row + 1, column=column + 1).value
                    case_value[key] = value
                cases_value[case_name] = case_value
    logging.debug("cases_value:%s" % cases_value)
    return cases_value


# 获取入参出参
def get_input_and_output(parameters, case_value):
    input_values = {}
    out_values = {}
    for parameter in parameters:
        if parameter[1].lower() == "in":
            ret = re.search("\${(.*)}", parameter[3])
            try:
                key = ret.group(1)
                if ".input." in key or ".out." in key.lower():
                    get_step_value = key.split(".")
                    value = "self.dict_" + get_step_value[0] + "." + get_step_value[2]
                else:
                    value = case_value.get(key)
                input_values[parameter[0]] = parameter[2] + "|" + value
            except:
                input_values[parameter[0]] = parameter[2] + "|" + parameter[3]
        elif parameter[1].lower() == "out":
            ret = re.search("\[(.*)\]", parameter[3])
            try:
                judge = ret.group(1)
            except:
                judge = "=="
            assert_judge = get_output_assert(judge.lower())

            ret = re.search("\${(.*)}", parameter[3])
            try:
                key = ret.group(1)
                if ".input." in key or ".out." in key.lower():
                    get_step_value = key.split(".")
                    value = "self.dict_" + get_step_value[0] + "." + get_step_value[2]
                else:
                    value = case_value.get(key)
                out_values[parameter[0]] = assert_judge + "|" + parameter[2] + "|" + value
            except:
                out_values[parameter[0]] = assert_judge + "|" + parameter[2] + "|" + parameter[3]
    return input_values, out_values


# 获取输出的断言
def get_output_assert(judge):
    if judge == "==":
        assert_judge = "assertEqual"
    elif judge == "!=":
        assert_judge = "assertNotEqual"
    elif judge == "in":
        assert_judge = "assertIn"
    elif judge == "not in":
        assert_judge = "assertNotIn"
    elif judge == "none":
        assert_judge = "assertIsNone"
    elif judge == "not none":
        assert_judge = "assertIsNotNone"
    return assert_judge


# 生成py文件
def create_py_test(case_file_name, case_template, cases_value):
    for case_name, case_value in cases_value.items():
        case = os.path.join(os.path.split(case_file_name)[0],
                            os.path.split(case_file_name)[1].split(".")[0] + "_" + case_name + ".py")
        logging.debug("case:%s" % case)
        logging.debug("case_name:%s" % case_name)
        logging.debug("case_value:%s" % case_value)
        with open(case, 'w', encoding='UTF-8') as f:
            f.write(
                "# -*- coding: utf-8 -*-\n"
                "from suite.douyin.projectresource.unittest_testcase import * \n\n"
                "class TestCase(OTestCase):\n"
                "\tdef testRun(self):\n")
            step_num = 0
            for step, parameters in case_template.items():
                step_num += 1
                logging.debug("step:%s" % step)
                logging.debug("parameters:%s" % parameters)
                f.write("\t\t#------------------------------%s---------------------------------\n"
                        "\t\ttry:\n"
                        "\t\t\tlog.logger.info('\\n' + 20 * '* ' + '%s' + 20 * ' *')\n"
                        "\t\t\tdict = {\n" % (step, step))

                input_values, out_values = get_input_and_output(parameters, case_value)
                for input_key, input_value in input_values.items():
                    if input_value.startswith("self") or input_value.split("|")[0].lower() == "int":
                        f.write("\t\t\t\t'%s': %s,\n" % (input_key, input_value.split("|")[1]))
                    else:
                        f.write("\t\t\t\t'%s': '%s',\n" % (input_key, input_value.split("|")[1]))

                function = str(step.split("|")[1])
                f.write("\t\t\t}\n"
                        "\t\t\tself.dict_step%s = self.%s(dict)\n" % (step_num, function))

                for out_key, out_value in out_values.items():
                    if out_value.split("|")[1].lower() == "int":
                        if out_value.startswith("self"):
                            f.write("\t\t\tself.%s(self.dict_step%s.get(%s), %s)\n" % (
                            out_value.split("|")[0], step_num, out_value.split("|")[2], out_key))
                        else:
                            f.write("\t\t\tself.%s(self.dict_step%s.get('%s'), %s)\n" % (
                            out_value.split("|")[0], step_num, out_value.split("|")[2], out_key))
                    elif out_value.split("|")[1].lower() == "str":
                        if out_value.startswith("self"):
                            f.write("\t\t\tself.%s(self.dict_step%s.get(%s), '%s')\n" % (
                            out_value.split("|")[0], step_num, out_value.split("|")[2], out_key))
                        else:
                            f.write("\t\t\tself.%s(self.dict_step%s.get('%s'), '%s')\n" % (
                            out_value.split("|")[0], step_num, out_value.split("|")[2], out_key))

                f.write("\t\t\tstatus = 'Success'\n"
                        "\t\texcept Exception as e:\n"
                        "\t\t\tstatus = 'Fail'\n"
                        "\t\t\tlog.logger.error(e)\n"
                        "\t\t\traise e\n"
                        "\t\tfinally:\n"
                        "\t\t\tlog.logger.info('\\n' + 20 * '* ' + '%s ' + status + 20 * ' *')\n" % step)
            f.write("if __name__ == '__main__':\n"
                    "\tunittest.main()")
        logging.info(case + " Sucess...")


if __name__ == '__main__':
    traverse_project()
